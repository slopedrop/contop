"""
Unit tests for core/document_tools.py - Document I/O tools.

Tests read_pdf, read_image, read_excel, write_excel with mocked libraries
and fixture files.

Module under test: core.document_tools
"""
import json
import os

import pytest

from core.document_tools import read_pdf, read_image, read_excel, write_excel


# --- read_pdf tests ---

@pytest.mark.unit
class TestReadPdf:
    async def test_read_nonexistent_pdf_returns_error(self, tmp_path):
        result = await read_pdf(str(tmp_path / "nope.pdf"))
        assert result["status"] == "error"

    async def test_read_pdf_missing_pymupdf_returns_error(self, tmp_path, monkeypatch):
        """If pymupdf4llm is not installed, return a helpful error."""
        p = tmp_path / "test.pdf"
        p.write_bytes(b"%PDF-1.4 test")

        import builtins
        real_import = builtins.__import__

        def mock_import(name, *args, **kwargs):
            if "pymupdf4llm" in name or "fitz" in name:
                raise ImportError("mocked")
            return real_import(name, *args, **kwargs)

        monkeypatch.setattr(builtins, "__import__", mock_import)
        result = await read_pdf(str(p))
        assert result["status"] == "error"


# --- read_image tests ---

@pytest.mark.unit
class TestReadImage:
    async def test_read_nonexistent_image_returns_error(self, tmp_path):
        result = await read_image(str(tmp_path / "nope.png"))
        assert result["status"] == "error"

    async def test_read_image_happy_path(self, tmp_path):
        """Create a minimal valid image and read it."""
        try:
            from PIL import Image
        except ImportError:
            pytest.skip("Pillow not installed")

        p = tmp_path / "test.png"
        img = Image.new("RGB", (100, 50), color="red")
        img.save(str(p))

        result = await read_image(str(p))
        assert result["status"] == "success"
        assert "image_b64" in result
        assert result["width"] == 100
        assert result["height"] == 50


# --- read_excel tests ---

@pytest.mark.unit
class TestReadExcel:
    async def test_read_nonexistent_excel_returns_error(self, tmp_path):
        result = await read_excel(str(tmp_path / "nope.xlsx"))
        assert result["status"] == "error"

    async def test_read_excel_happy_path(self, tmp_path):
        """Create a minimal xlsx and read it."""
        try:
            import openpyxl
        except ImportError:
            pytest.skip("openpyxl not installed")

        p = tmp_path / "test.xlsx"
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Data"
        ws["A1"] = "Name"
        ws["B1"] = "Value"
        ws["A2"] = "Alice"
        ws["B2"] = 42
        wb.save(str(p))

        result = await read_excel(str(p))
        assert result["status"] == "success"
        assert "Name" in result["content"]
        assert "Alice" in result["content"]
        assert "Data" in result["sheets_available"]


# --- write_excel tests ---

@pytest.mark.unit
class TestWriteExcel:
    async def test_write_excel_set_cell(self, tmp_path):
        try:
            import openpyxl
        except ImportError:
            pytest.skip("openpyxl not installed")

        import json
        p = tmp_path / "output.xlsx"
        ops = json.dumps([{"action": "set_cell", "cell": "A1", "value": "Hello"}])
        result = await write_excel(str(p), operations=ops)
        assert result["status"] == "success"

        # Verify the cell was written
        wb = openpyxl.load_workbook(str(p))
        assert wb.active["A1"].value == "Hello"

    async def test_write_excel_no_operations_returns_error(self, tmp_path):
        import json
        p = tmp_path / "output.xlsx"
        result = await write_excel(str(p), operations=json.dumps([]))
        assert result["status"] == "error" or result.get("operations_applied") == 0
