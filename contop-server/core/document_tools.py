"""
Document tools — structured document reading and writing for the execution agent.

Provides read_pdf, read_image, read_excel, write_excel as ADK FunctionTools.
These give the agent native document access without CLI workarounds.

All tools follow the standard async pattern: async def, dict return with
status field, logger.info at entry, try/except with logger.exception.
"""
import asyncio
import json
import logging
import time as _time
from pathlib import Path

logger = logging.getLogger(__name__)


async def read_pdf(file_path: str, pages: str = "") -> dict:
    """Read a PDF file and return its content as markdown.

    Args:
        file_path: Absolute path to the PDF file.
        pages: Page range (e.g., "1-5", "3", "10-20"). Default: first 20 pages.

    Returns dict with status, content, page_count, images.
    """
    logger.info("read_pdf called: file_path=%s, pages=%s", file_path, pages)
    start = _time.monotonic()
    try:
        import pymupdf4llm

        p = Path(file_path)
        if not p.exists():
            return {
                "status": "error",
                "description": f"File not found: {file_path}",
                "duration_ms": int((_time.monotonic() - start) * 1000),
                "voice_message": f"I couldn't find the PDF file {p.name}.",
            }

        # Parse page range
        page_list = None
        if pages:
            page_list = _parse_page_range(pages)

        def _read():
            md = pymupdf4llm.to_markdown(
                str(p),
                pages=page_list,
            )
            return md

        content = await asyncio.to_thread(_read)

        # Count pages from the file
        import pymupdf
        doc = pymupdf.open(str(p))
        total_pages = len(doc)
        doc.close()

        return {
            "status": "success",
            "content": content[:200_000],  # Cap content to avoid context explosion
            "page_count": total_pages,
            "pages_read": pages or f"1-{min(20, total_pages)}",
            "duration_ms": int((_time.monotonic() - start) * 1000),
        }
    except ImportError as exc:
        logger.error("read_pdf missing dependency: %s", exc)
        return {
            "status": "error",
            "description": f"Server dependency missing: {exc}. Restart the server after installing.",
            "duration_ms": int((_time.monotonic() - start) * 1000),
            "voice_message": "A required library is missing from the server.",
        }
    except Exception as exc:
        logger.exception("read_pdf failed")
        return {
            "status": "error",
            "description": str(exc),
            "duration_ms": int((_time.monotonic() - start) * 1000),
            "voice_message": "I had trouble reading that PDF.",
        }


def _parse_page_range(pages: str) -> list[int]:
    """Parse page range string like '1-5' or '3' into a list of 0-based page indices."""
    result = []
    for part in pages.split(","):
        part = part.strip()
        if "-" in part:
            s, e = part.split("-", 1)
            start = max(0, int(s.strip()) - 1)
            end = int(e.strip())
            result.extend(range(start, end))
        else:
            result.append(int(part.strip()) - 1)
    return result


async def read_image(file_path: str) -> dict:
    """Read an image file and return it as base64 JPEG.

    Resizes images larger than 2000px on either dimension.

    Args:
        file_path: Absolute path to the image file.

    Returns dict with status, image_b64, width, height, mime_type.
    """
    logger.info("read_image called: file_path=%s", file_path)
    start = _time.monotonic()
    try:
        from PIL import Image
        import base64
        import io

        p = Path(file_path)
        if not p.exists():
            return {
                "status": "error",
                "description": f"File not found: {file_path}",
                "duration_ms": int((_time.monotonic() - start) * 1000),
                "voice_message": f"I couldn't find the image file {p.name}.",
            }

        def _read():
            img = Image.open(str(p))
            # Resize if too large
            max_dim = 2000
            if img.width > max_dim or img.height > max_dim:
                ratio = min(max_dim / img.width, max_dim / img.height)
                new_size = (int(img.width * ratio), int(img.height * ratio))
                img = img.resize(new_size, Image.LANCZOS)

            # Convert to JPEG
            if img.mode in ("RGBA", "LA", "P"):
                img = img.convert("RGB")
            buf = io.BytesIO()
            img.save(buf, format="JPEG", quality=85)
            return base64.b64encode(buf.getvalue()).decode(), img.width, img.height

        b64_data, w, h = await asyncio.to_thread(_read)

        return {
            "status": "success",
            "image_b64": b64_data,
            "width": w,
            "height": h,
            "mime_type": "image/jpeg",
            "duration_ms": int((_time.monotonic() - start) * 1000),
        }
    except ImportError as exc:
        logger.error("read_image missing dependency: %s", exc)
        return {
            "status": "error",
            "description": f"Server dependency missing: {exc}. Restart the server after installing.",
            "duration_ms": int((_time.monotonic() - start) * 1000),
            "voice_message": "A required library is missing from the server.",
        }
    except Exception as exc:
        logger.exception("read_image failed")
        return {
            "status": "error",
            "description": str(exc),
            "duration_ms": int((_time.monotonic() - start) * 1000),
            "voice_message": "I had trouble reading that image.",
        }


async def read_excel(file_path: str, sheet: str = "", cell_range: str = "") -> dict:
    """Read an Excel (.xlsx) file and return content as a markdown table.

    Args:
        file_path: Absolute path to the Excel file.
        sheet: Sheet name or index (default: first sheet).
        cell_range: Cell range like "A1:F50" (default: first 100 rows, all columns).

    Returns dict with status, content, sheet_name, rows, columns, sheets_available.
    """
    logger.info("read_excel called: file_path=%s, sheet=%s, cell_range=%s", file_path, sheet, cell_range)
    start = _time.monotonic()
    try:
        import openpyxl

        p = Path(file_path)
        if not p.exists():
            return {
                "status": "error",
                "description": f"File not found: {file_path}",
                "duration_ms": int((_time.monotonic() - start) * 1000),
                "voice_message": f"I couldn't find the Excel file {p.name}.",
            }

        def _read():
            wb = openpyxl.load_workbook(str(p), data_only=True)
            sheet_names = wb.sheetnames

            if sheet:
                if sheet.isdigit():
                    ws = wb.worksheets[int(sheet)]
                else:
                    ws = wb[sheet]
            else:
                ws = wb.active or wb.worksheets[0]

            sheet_name = ws.title

            if cell_range:
                rows = list(ws[cell_range])
            else:
                rows = list(ws.iter_rows(max_row=100))

            # Truncate at 200 rows
            rows = rows[:200]

            # Convert to markdown table
            if not rows:
                wb.close()
                return "", sheet_name, 0, 0, sheet_names

            # Build header from first row
            header = [str(cell.value or "") for cell in rows[0]]
            separator = ["-" * max(3, len(h)) for h in header]
            lines = [
                "| " + " | ".join(header) + " |",
                "| " + " | ".join(separator) + " |",
            ]
            for row in rows[1:]:
                vals = [str(cell.value or "") for cell in row]
                # Pad if fewer columns
                while len(vals) < len(header):
                    vals.append("")
                lines.append("| " + " | ".join(vals[:len(header)]) + " |")

            wb.close()
            return "\n".join(lines), sheet_name, len(rows), len(header), sheet_names

        content, sheet_name, row_count, col_count, sheets = await asyncio.to_thread(_read)

        return {
            "status": "success",
            "content": content,
            "sheet_name": sheet_name,
            "rows": row_count,
            "columns": col_count,
            "sheets_available": sheets,
            "duration_ms": int((_time.monotonic() - start) * 1000),
        }
    except ImportError as exc:
        logger.error("read_excel missing dependency: %s", exc)
        return {
            "status": "error",
            "description": f"Server dependency missing: {exc}. Restart the server after installing.",
            "duration_ms": int((_time.monotonic() - start) * 1000),
            "voice_message": "A required library is missing from the server.",
        }
    except Exception as exc:
        logger.exception("read_excel failed: %s", exc)
        return {
            "status": "error",
            "description": str(exc),
            "duration_ms": int((_time.monotonic() - start) * 1000),
            "voice_message": "I had trouble reading that Excel file.",
        }


async def write_excel(file_path: str, operations: str) -> dict:
    """Write to an Excel (.xlsx) file using structured operations.

    Args:
        file_path: Absolute path to the Excel file (creates if doesn't exist).
        operations: JSON string of operations, e.g.:
            [{"action": "set_cell", "cell": "A1", "value": "Hello"},
             {"action": "set_style", "cell": "A1", "bold": true},
             {"action": "merge_cells", "range": "A1:C1"},
             {"action": "set_column_width", "column": "A", "width": 20},
             {"action": "add_sheet", "name": "Sheet2"}]

    Returns dict with status, operations_applied, file_path.
    """
    logger.info("write_excel called: file_path=%s", file_path)
    start = _time.monotonic()
    try:
        import openpyxl
        from openpyxl.styles import Font

        p = Path(file_path)
        ops = json.loads(operations)

        def _write():
            if p.exists():
                wb = openpyxl.load_workbook(str(p))
            else:
                wb = openpyxl.Workbook()

            ws = wb.active
            applied = 0

            for op in ops:
                action = op.get("action", "")
                if action == "set_cell":
                    cell = op.get("cell", "A1")
                    value = op.get("value", "")
                    sheet_name = op.get("sheet")
                    target = wb[sheet_name] if sheet_name and sheet_name in wb.sheetnames else ws
                    target[cell] = value
                    applied += 1
                elif action == "set_style":
                    cell = op.get("cell", "A1")
                    sheet_name = op.get("sheet")
                    target = wb[sheet_name] if sheet_name and sheet_name in wb.sheetnames else ws
                    font_kw = {}
                    if "bold" in op:
                        font_kw["bold"] = op["bold"]
                    if "italic" in op:
                        font_kw["italic"] = op["italic"]
                    if "font_size" in op:
                        font_kw["size"] = op["font_size"]
                    if font_kw:
                        target[cell].font = Font(**font_kw)
                    applied += 1
                elif action == "merge_cells":
                    cell_range = op.get("range", "")
                    ws.merge_cells(cell_range)
                    applied += 1
                elif action == "set_column_width":
                    column = op.get("column", "A")
                    width = op.get("width", 15)
                    ws.column_dimensions[column].width = width
                    applied += 1
                elif action == "add_sheet":
                    name = op.get("name", "Sheet")
                    wb.create_sheet(title=name)
                    applied += 1

            # Atomic write: temp file then rename
            tmp = p.with_suffix(".xlsx.tmp")
            wb.save(str(tmp))
            wb.close()
            tmp.replace(p)
            return applied

        applied = await asyncio.to_thread(_write)

        return {
            "status": "success",
            "operations_applied": applied,
            "file_path": file_path,
            "duration_ms": int((_time.monotonic() - start) * 1000),
        }
    except ImportError as exc:
        logger.error("write_excel missing dependency: %s", exc)
        return {
            "status": "error",
            "description": f"Server dependency missing: {exc}. Restart the server after installing.",
            "duration_ms": int((_time.monotonic() - start) * 1000),
            "voice_message": "A required library is missing from the server.",
        }
    except json.JSONDecodeError as exc:
        return {
            "status": "error",
            "description": f"Invalid JSON in operations: {exc}",
            "duration_ms": int((_time.monotonic() - start) * 1000),
            "voice_message": "The operations format was invalid.",
        }
    except Exception as exc:
        logger.exception("write_excel failed")
        return {
            "status": "error",
            "description": str(exc),
            "duration_ms": int((_time.monotonic() - start) * 1000),
            "voice_message": "I had trouble writing to the Excel file.",
        }
