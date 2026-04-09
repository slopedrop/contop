"""
Document conversion tool - platform-agnostic conversion between file formats.

Registered as an agent FunctionTool when the office-documents skill is enabled.
Detects available tools on the system and uses the best method automatically.
"""
import asyncio
import logging
import os
import platform
import shutil
import subprocess
import time as _time
from pathlib import Path

logger = logging.getLogger(__name__)

_PLATFORM = platform.system()

# Supported conversions: (input_ext, output_format) → handler
# Handlers are tried in order; first available one wins.
_LIBREOFFICE_FORMATS = {"pdf", "csv", "html", "png", "jpg"}
_OFFICE_COM_FORMATS = {"pdf"}


def _find_libreoffice() -> str | None:
    """Return the path to LibreOffice soffice binary, or None."""
    candidates = []
    if _PLATFORM == "Windows":
        for base in [
            os.environ.get("PROGRAMFILES", r"C:\Program Files"),
            os.environ.get("PROGRAMFILES(X86)", r"C:\Program Files (x86)"),
        ]:
            if base:
                candidates.append(Path(base) / "LibreOffice" / "program" / "soffice.exe")
    elif _PLATFORM == "Darwin":
        candidates.append(Path("/Applications/LibreOffice.app/Contents/MacOS/soffice"))
    else:
        # Linux: check PATH
        which = shutil.which("libreoffice") or shutil.which("soffice")
        if which:
            return which

    for c in candidates:
        if c.exists():
            return str(c)

    # Final fallback: check PATH
    return shutil.which("libreoffice") or shutil.which("soffice")


async def _has_office_com() -> bool:
    """Check if Microsoft Office COM automation is available (Windows only)."""
    if _PLATFORM != "Windows":
        return False
    try:
        proc = await asyncio.create_subprocess_exec(
            "powershell", "-NoProfile", "-Command",
            "try { New-Object -ComObject Excel.Application | Out-Null; $true } catch { $false }",
            stdout=asyncio.subprocess.PIPE,
            stderr=asyncio.subprocess.DEVNULL,
        )
        stdout, _ = await asyncio.wait_for(proc.communicate(), timeout=10)
        return stdout.decode("utf-8", errors="replace").strip().lower() == "true"
    except Exception:
        return False


async def _convert_via_office_com(
    input_path: Path, output_path: Path, output_format: str,
    sheet_name: str = "",
) -> dict:
    """Convert using Microsoft Office COM automation (Windows only)."""
    ext = input_path.suffix.lower()

    if ext in (".xlsx", ".xls", ".xlsm"):
        # Excel COM
        sheet_clause = ""
        if sheet_name:
            sheet_clause = f"""
  $ws = $wb.Sheets.Item('{sheet_name}')
  $ws.Activate()"""

        ps_script = f"""
$excel = New-Object -ComObject Excel.Application
$excel.Visible = $false
$excel.DisplayAlerts = $false
try {{
  $wb = $excel.Workbooks.Open('{input_path}'){sheet_clause}
  $wb.ExportAsFixedFormat(0, '{output_path}')
  $wb.Close($false)
  Write-Output 'OK'
}} finally {{
  $excel.Quit()
  [System.Runtime.InteropServices.Marshal]::ReleaseComObject($excel) | Out-Null
}}
"""
    elif ext in (".docx", ".doc"):
        ps_script = f"""
$word = New-Object -ComObject Word.Application
$word.Visible = $false
try {{
  $doc = $word.Documents.Open('{input_path}')
  $doc.ExportAsFixedFormat('{output_path}', 17)
  $doc.Close($false)
  Write-Output 'OK'
}} finally {{
  $word.Quit()
  [System.Runtime.InteropServices.Marshal]::ReleaseComObject($word) | Out-Null
}}
"""
    elif ext in (".pptx", ".ppt"):
        ps_script = f"""
$ppt = New-Object -ComObject PowerPoint.Application
try {{
  $pres = $ppt.Presentations.Open('{input_path}', $true, $false, $false)
  $pres.ExportAsFixedFormat('{output_path}', 2)
  $pres.Close()
  Write-Output 'OK'
}} finally {{
  $ppt.Quit()
  [System.Runtime.InteropServices.Marshal]::ReleaseComObject($ppt) | Out-Null
}}
"""
    else:
        return {"status": "error", "description": f"Office COM does not support {ext} files."}

    proc = await asyncio.create_subprocess_exec(
        "powershell", "-NoProfile", "-Command", ps_script,
        stdout=asyncio.subprocess.PIPE,
        stderr=asyncio.subprocess.PIPE,
    )
    stdout, stderr = await asyncio.wait_for(proc.communicate(), timeout=60)
    stdout_str = stdout.decode("utf-8", errors="replace").strip()
    stderr_str = stderr.decode("utf-8", errors="replace").strip()

    if proc.returncode == 0 and "OK" in stdout_str:
        return {"status": "success"}
    return {
        "status": "error",
        "description": f"Office COM failed (exit {proc.returncode}): {stderr_str or stdout_str}",
    }


async def _convert_via_libreoffice(
    soffice: str, input_path: Path, output_dir: Path, output_format: str,
) -> dict:
    """Convert using LibreOffice headless mode."""
    proc = await asyncio.create_subprocess_exec(
        soffice, "--headless", "--convert-to", output_format,
        "--outdir", str(output_dir), str(input_path),
        stdout=asyncio.subprocess.PIPE,
        stderr=asyncio.subprocess.PIPE,
    )
    stdout, stderr = await asyncio.wait_for(proc.communicate(), timeout=120)
    stdout_str = stdout.decode("utf-8", errors="replace").strip()

    if proc.returncode == 0:
        return {"status": "success"}
    return {
        "status": "error",
        "description": f"LibreOffice failed (exit {proc.returncode}): {stdout_str}",
    }


async def _convert_via_python(
    input_path: Path, output_path: Path, output_format: str,
    sheet_name: str = "",
) -> dict:
    """Convert using pure Python libraries (pip-installable fallback)."""

    def _do_convert() -> dict:
        ext = input_path.suffix.lower()

        if ext in (".xlsx", ".xls", ".xlsm") and output_format == "csv":
            import openpyxl
            wb = openpyxl.load_workbook(str(input_path), data_only=True)
            ws = wb[sheet_name] if sheet_name and sheet_name in wb.sheetnames else wb.active
            import csv
            with open(str(output_path), "w", newline="", encoding="utf-8") as f:
                writer = csv.writer(f)
                for row in ws.iter_rows(values_only=True):
                    writer.writerow(row)
            wb.close()
            return {"status": "success"}

        if ext in (".xlsx", ".xls", ".xlsm") and output_format == "pdf":
            try:
                from fpdf import FPDF
            except ImportError:
                return {
                    "status": "error",
                    "description": "No PDF converter available. Install fpdf2: pip install fpdf2",
                    "install_hint": "pip install fpdf2",
                }

            import openpyxl
            wb = openpyxl.load_workbook(str(input_path), data_only=True)
            ws = wb[sheet_name] if sheet_name and sheet_name in wb.sheetnames else wb.active

            pdf = FPDF()
            pdf.add_page(orientation="L")
            pdf.set_auto_page_break(auto=True, margin=15)
            pdf.set_font("Helvetica", size=8)

            for row in ws.iter_rows(values_only=True):
                line = "  |  ".join(str(c) if c is not None else "" for c in row)
                pdf.cell(0, 6, line, new_x="LMARGIN", new_y="NEXT")

            pdf.output(str(output_path))
            wb.close()
            return {"status": "success"}

        return {
            "status": "error",
            "description": f"No Python handler for {ext} -> {output_format}.",
        }

    return await asyncio.to_thread(_do_convert)


async def convert_document(
    input_path: str,
    output_format: str = "pdf",
    output_path: str = "",
    sheet_name: str = "",
) -> dict:
    """Convert a document to another format (PDF, CSV, PNG, HTML, etc.).

    Automatically detects the best conversion method available on this system:
    Microsoft Office COM (Windows), LibreOffice (all platforms), or Python
    libraries (fallback). Install missing packages with pip if needed.

    Args:
        input_path: Absolute path to the source document (.xlsx, .docx, .pptx, .csv, .html, etc.).
        output_format: Target format: "pdf", "csv", "png", "jpg", "html". Default "pdf".
        output_path: Where to save the result. If empty, saves next to input file with new extension.
        sheet_name: For spreadsheets, which sheet to convert. Empty = active/first sheet.

    Returns:
        dict with status, output_path, method used, and duration_ms.
    """
    logger.info(
        "convert_document called: input=%s, format=%s, output=%s, sheet=%s",
        input_path, output_format, output_path, sheet_name,
    )
    start = _time.monotonic()
    output_format = output_format.lower().lstrip(".")

    try:
        inp = Path(input_path)
        if not inp.exists():
            return {
                "status": "error",
                "description": f"Input file not found: {input_path}",
                "duration_ms": int((_time.monotonic() - start) * 1000),
                "voice_message": "That file doesn't exist.",
            }

        # Determine output path
        if output_path:
            out = Path(output_path)
        else:
            out = inp.with_suffix(f".{output_format}")

        out.parent.mkdir(parents=True, exist_ok=True)

        method_used = ""

        # Strategy 1: Office COM (Windows only, PDF output, Office formats)
        if (
            _PLATFORM == "Windows"
            and output_format in _OFFICE_COM_FORMATS
            and inp.suffix.lower() in (
                ".xlsx", ".xls", ".xlsm",
                ".docx", ".doc",
                ".pptx", ".ppt",
            )
        ):
            if await _has_office_com():
                result = await _convert_via_office_com(inp, out, output_format, sheet_name)
                if result["status"] == "success":
                    method_used = "office_com"

        # Strategy 2: LibreOffice headless (all platforms, many formats)
        if not method_used:
            soffice = _find_libreoffice()
            if soffice and output_format in _LIBREOFFICE_FORMATS:
                result = await _convert_via_libreoffice(soffice, inp, out.parent, output_format)
                if result["status"] == "success":
                    # LibreOffice names the output after the input file
                    lo_output = out.parent / f"{inp.stem}.{output_format}"
                    if lo_output != out and lo_output.exists():
                        lo_output.rename(out)
                    method_used = "libreoffice"

        # Strategy 3: Python libraries (pip-installable fallback)
        if not method_used:
            result = await _convert_via_python(inp, out, output_format, sheet_name)

            if result["status"] == "success":
                method_used = "python"
            elif "install_hint" in result:
                return {
                    "status": "error",
                    "description": result["description"],
                    "install_hint": result["install_hint"],
                    "duration_ms": int((_time.monotonic() - start) * 1000),
                    "voice_message": f"I need an extra package to convert to {output_format}. I can install it with pip.",
                }

        if not method_used:
            available = []
            if _PLATFORM == "Windows":
                available.append("Microsoft Office (for COM automation)")
            available.append("LibreOffice (libreoffice --headless)")
            available.append("Python fpdf2 (pip install fpdf2)")
            return {
                "status": "error",
                "description": (
                    f"No conversion method available for {inp.suffix} -> {output_format}. "
                    f"Install one of: {'; '.join(available)}"
                ),
                "duration_ms": int((_time.monotonic() - start) * 1000),
                "voice_message": f"I don't have the tools to convert to {output_format}. Check the error for install options.",
            }

        return {
            "status": "success",
            "output_path": str(out),
            "method": method_used,
            "input_format": inp.suffix.lstrip("."),
            "output_format": output_format,
            "duration_ms": int((_time.monotonic() - start) * 1000),
            "voice_message": f"Converted to {output_format} successfully.",
        }

    except Exception as exc:
        logger.exception("convert_document failed")
        return {
            "status": "error",
            "description": str(exc),
            "duration_ms": int((_time.monotonic() - start) * 1000),
            "voice_message": f"The document conversion failed.",
        }
